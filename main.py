import time
from selenium import webdriver
from bs4 import BeautifulSoup
from tkinter import *
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import os

class Linkedin():
    def __init__(self):
        self.config = self.read_config()

    def read_config(self):
        config = {}
        try:
            with open('config.txt', 'r') as f:
                for line in f:
                    key, value = line.strip().split(' ')
                    config[key.strip()] = value.strip()
        except FileNotFoundError:
            print("Config file not found. Please create a config.txt file.")
            exit(1)
        except ValueError:
            print("Invalid config file format. Each line should be in the format: key = value")
            exit(1)
        return config

    def getData(self):
        driver = webdriver.Safari()
        driver.get('https://www.linkedin.com/login')
        driver.find_element(By.ID, 'username').send_keys(self.config['username'])
        driver.find_element(By.ID, 'password').send_keys(self.config['password'])
        driver.find_element(By.XPATH, "//*[@type='submit']").click()

        global data
        data = []

        for no in range(1, 100):
            search_url = self.config['search_url'].format(no)
            driver.get(search_url)
            driver.maximize_window()
            print("Going to scrape Page {} data".format(no))
            for scroll in range(2):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)

            search = BeautifulSoup(driver.page_source, 'lxml')
            peoples = search.findAll('div', class_='mb1')  # Find all profile containers
            count = 0
            
            for people in peoples:
                count += 1

                # Extract the profile link and name
                profile_link = people.find('a', class_='app-aware-link')
                if profile_link:
                    profile_url = profile_link['href'].split('?')[0]
                    full_name = profile_link.get_text(strip=True)  # Get the full text

                    # Remove unwanted text
                    name = full_name.split('View')[0].strip()  # Split and take the first part

                    # Extract the location
                    location_div = people.find('div', class_='entity-result__secondary-subtitle')
                    location = location_div.get_text(strip=True) if location_div else 'None'  # Get the location text

                    # Append the extracted data to the data list
                    data.append({'profile_url': profile_url, 'name': name, 'location': location})
            print("!!!!!! Data scrapped !!!!!!")

            # Write data to Excel after each page
            self.writeData(data)

        driver.quit()

    def writeData(self, data):
        file_name = "linkedin-search-data.xlsx"
        
        # Check if the file exists
        if os.path.exists(file_name):
            # Load the existing workbook
            workbook = load_workbook(file_name)
            worksheet = workbook['Peoples']
        else:
            # Create a new workbook and worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Peoples'
            # Write headers only if the file is new
            worksheet.append(['profile_url', 'Name', 'Location'])  # Write headers

        # Write the data to the worksheet
        for entry in data:
            worksheet.append([entry['profile_url'], entry['name'], entry['location']])

        # Save the workbook
        workbook.save(file_name)
        workbook.close()

    def start(self):
        self.getData()

if __name__ == "__main__":
    obJH = Linkedin()
    obJH.start()
